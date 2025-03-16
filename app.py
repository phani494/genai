
from flask import Flask, render_template, request, jsonify
import os
from openai import OpenAI
from PyPDF2 import PdfReader
import openpyxl
#from docx import Document  # For reading Word files
#from pptx import Presentation  # For reading PowerPoint files
#import textract

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = './uploads'


def extract_text_from_file(file_path):
    """Extract text from PDF, Excel, Word, PowerPoint, or text files."""
    _, file_extension = os.path.splitext(file_path)
    
    # Handle PDF files
    if file_extension == ".pdf":
        reader = PdfReader(file_path)
        text = "".join([page.extract_text() for page in reader.pages])
        return text
    
    # Handle Excel files
    elif file_extension in [".xls", ".xlsx"]:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                row_text = ", ".join([str(cell) if cell is not None else "" for cell in row])
                text.append(row_text)
        return "\n".join(text)
    
    # Handle Word files
    elif file_extension == ".docx":
        document = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in document.paragraphs])
        return text
    
    # Handle PowerPoint files
    elif file_extension == ".pptx":
        presentation = Presentation(file_path)
        text = []
        for slide in presentation.slides:
            for shape in slide.shapes:
                if shape.has_text_frame:
                    text.append(shape.text)
        return "\n".join(text)
    
    # Handle plain text files
    elif file_extension == ".txt":
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    # Unsupported file format
    else:
        return "Unsupported file format."


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file uploads and process the content."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    uploaded_file = request.files['file']
    if uploaded_file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
    uploaded_file.save(file_path)

    file_content = extract_text_from_file(file_path)
    return jsonify({"content": file_content})


@app.route('/ask', methods=['POST'])
def ask_question():
    """Handle user questions."""
    data = request.get_json()
    question = data.get("question")
    file_content = data.get("content")
    instruction = file_content + "\n\nFrom the above text answer the question: " + question
    
    if not question or not file_content:
        return jsonify({"error": "Invalid input"}), 400

    try:
        client = OpenAI(api_key="sk-Dz5KTQ8v8QLuucdhQHZQcz1ujxCQy2ZzyUd-fZwIEGT3BlbkFJTP4wGuY4xY8UQVntn2dOvS6FdpbDwWWA0ILpiicgoA")
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{
                "role": "user",
                "content": instruction
            }],
            temperature=1,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0, n=1
        )
        return jsonify({"answer": response.choices[0].message.content})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True)
